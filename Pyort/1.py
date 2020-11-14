"""
#---------------------------------------------------------------
"""


#Hoşgeldiniz Lütfen Üstteki Run kısmından Run Module'ü seçin

#   Kod çalışınca 'Çıktı' isimli bir dosya oluşacak, o dosyanın 2. sayfasında
#   Grafiklerinizi görebilirsiniz ve 'Çıktı' isimli dosyayı dilediğiniz gibi
#   düzenleyebilirsiniz.




#Sık Yapılan Hatalar!

#1- '1' isimli excel dosyası açık iken kodu çalıştırırsanız hata verir!
#   Böyle bir durumda excel dosyasın kapatıp tekrar Run Module'ü seçin.

#2- '1' isimli excel dosyası ile '1' isimli kodun aynı dosyada olduğundan emin olun!

#3- Kod üzerinde herhangi bir değişiklik yapmayın bu durum kodu bozabilir!

#4- Halen bir sorun yaşıyorsanız lütfen destek için iletişime geçin.

#Ayarlar:
SıfırKuralı = True
BarGrafiği = True
X_GrafiğiBaşlık = "Sene"
Y_GrafiğiBaşlık = "Net"

#İletişim:

#Gökdeniz Kaymak
#BotE Software
#gokdenizk.be@gmail.com

"""
#---------------------------------------------------------------
"""

import  openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series, LineChart, ScatterChart

excel = openpyxl.load_workbook("1.xlsx")
ws = excel.active
Aralıklar = []

#--------------------------------Fonksiyonlar--------------
def SıfırMı(char):
    if SıfırKuralı :
        if char == "0":
            return False
        else:
            return True
    else:
        return True
#-------------------------------------Veri Analizi--------------------
x = 2
y = 1
root = False
new = []
while str(ws.cell(y, x).value).replace(" ", "") != "x":

    if ws.cell(y,x).value != None and ws.cell(y,x).value != "x" and root == False:
        new.append(ws.cell(y,x).value)
        new.append(x)
        root = True

    elif ws.cell(y,x).value != None and ws.cell(y,x).value != "x" and root == True:
        new.append(x-1)
        Aralıklar.append(new)
        new = []
        new = [ws.cell(y, x).value,x]
    x += 1
new.append(x-1)
Aralıklar.append(new)
new = []

Sheet = excel.worksheets[1]


kişiSayısı = 0
Toplam = 0
Sonuç = 0


for Ders in Aralıklar:
    x = Ders[1]
    y = 3
    Sheet.cell(1, x).value = Ders[0]
    while x <= Ders[2]:
        Sheet.cell(2, x).value = ws.cell(2, x).value
        while y < 900:
            if ws.cell(y, x).value != None and SıfırMı(str(ws.cell(y, x).value).replace(" ", "")):
                Toplam += float(ws.cell(y, x).value)
                kişiSayısı += 1

            y += 1
        Sonuç = Toplam / kişiSayısı
        Sheet.cell(3, x).value = round(Sonuç, 2)
        Sonuç = 0
        Toplam = 0
        kişiSayısı = 0
        x += 1
        y = 3

#--------------------------------Grafikler-----------------
col = 4


for Ders in Aralıklar:
    for x in range(Ders[1], Ders[2]+1):
        Sheet.cell(x - Ders[1] + 3, Ders[1]).value = Sheet.cell(3, x).value
    for x in range(Ders[1], Ders[2] + 1):
        Sheet.cell(x - Ders[1] + 3, Ders[1] + 1).value = Sheet.cell(2, x).value

Pos = "A15"
number = 1
numb2 = 15
Alfa = [[1,"A"],[2,"B"],[3,"C"],[4,"D"],[5,"E"],[6,"F"],[7,"G"],[8,"H"],[9,"I"],[10,"J"],[11,"K"],[12,"L"],[13,"M"],[14,"N"],[15,"O"],[16,"P"],[17,"Q"],[18,"R"],[19,"S"],[20,"T"],[21,"U"],[22,"V"],[23,"W"],[24,"X"],[25,"Y"],[26,"Z"]]

for Ders in Aralıklar:
    if BarGrafiği == False:
        values = Reference(Sheet, min_col=Ders[1], min_row=3, max_row=Ders[2] - Ders[1] + 3)
        Dates = Reference(Sheet, min_col=Ders[1] + 1, min_row=3, max_row=Ders[2] - Ders[1] + 3)
        # chart = LineChart()
        chart = LineChart()
        Sheet.add_chart(chart, Pos)
        chart.title = Ders[0]
        chart.y_axis.title = Y_GrafiğiBaşlık
        chart.x_axis.title = X_GrafiğiBaşlık
        chart.add_data(values)
        s1 = chart.series[0]
        s1.marker.symbol = "triangle"
        chart.set_categories(Dates)
        if number + 10 < 26:
            number += 10
        else:
            number = 1
            numb2 += 15
        Pos = str(Alfa[number - 1][1]) + str(numb2)

    elif BarGrafiği:
        Sheet.cell(2, Ders[1]).value = Ders[0]
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = Ders[0]
        chart1.y_axis.title = Y_GrafiğiBaşlık
        chart1.x_axis.title = X_GrafiğiBaşlık

        data = Reference(Sheet, min_col=Ders[1], min_row=2, max_row=Ders[2] - Ders[1] + 3)
        cats = Reference(Sheet, min_col=Ders[1] + 1, min_row=3, max_row=Ders[2] - Ders[1] + 4)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4
        Sheet.add_chart(chart1, Pos)
        if number + 10 < 26:
            number += 10
        else:
            number = 1
            numb2 += 15
        Pos = str(Alfa[number - 1][1]) + str(numb2)

print("------------------------İşleminiz Başarı ile gerçekleşti------------------------")
excel.save("Çıktı.xlsx")





